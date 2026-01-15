import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urlparse, urljoin
import time
import pandas as pd
import zipfile
import os
import xml.etree.ElementTree as ET
import gzip
import io
from datetime import datetime

# Email Imports
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import smtplib

# Selenium & Excel Imports
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Webdriver Manager (Optional dependency management)
try:
    from webdriver_manager.chrome import ChromeDriverManager
    WEBDRIVER_MANAGER_AVAILABLE = True
except ImportError:
    WEBDRIVER_MANAGER_AVAILABLE = False

# --- CONFIGURATION ---
# Email Config
smtp_server = 'smtp.gmail.com'
smtp_port = 587
smtp_username = 'hemendra.rana@deptagency.com'
smtp_password = 'azvp nyjw leel rtdt'
sender_email = 'hemendra.rana@deptagency.com'

# List of emails for report recipients
recipients_birla = ["hariom.singh@deptagency.com", "anand.tigga@adityabirla.com",
    "avishek.kumar@adityabirla.com",
    "ramandeep.singh@adityabirla.com",
    "Aastha.narula@adityabirla.com",
    "Juhi.Parmar@adityabirla.com",
    "surbhi.gupta@adityabirla.com",
    "bhumika.khunt@deptagency.com",
    "shubham.rogye@deptagency.com",
    "gaurang.kapadia@deptagency.com",
    "shadab.sayed@deptagency.com",
    "ashita.choudhury@deptagency.com",
    "dhawal.mhatre@deptagency.com",
    "monica.ledwani@deptagency.com"]

# Define Sitemap URL
birla_opus_sitemap_url = 'https://www.birlaopus.com/sitemap.xml'

# Generate timestamp for report naming
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
date_string = datetime.now().strftime('%d-%b-%Y %I:%M %p')
output_dir_birla = f'birla_opus_performance_reports_{timestamp}'

# --- URL FILTERING FUNCTION ---
def filter_urls_ending_with_slash(urls):
    """
    Filter URLs to only include:
    1. Root domain URL (0 segments): https://www.birlaopus.com/
    2. Single-segment path URLs (1 segment): https://www.birlaopus.com/colour-catalogue/
    
    Excludes URLs with multiple path segments like:
    - https://www.birlaopus.com/colour-catalogue/something-else/
    - https://www.birlaopus.com/products/doors/interior/
    """
    from urllib.parse import urlparse
    
    filtered_urls = []
    zero_segment_count = 0
    one_segment_count = 0
    skipped_count = 0
    
    print("\n=== URL FILTERING DEBUG ===")
    
    for url in urls:
        parsed = urlparse(url)
        path = parsed.path
        
        # Skip non-HTML files (images, PDFs, etc.)
        non_html_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.pdf', '.xml', '.css', '.js', '.svg', '.webp', '.ico', '.zip', '.mp4', '.mp3', '.avi', '.mov']
        if any(path.lower().endswith(ext) for ext in non_html_extensions):
            skipped_count += 1
            continue
        
        # Skip URLs with query parameters or fragments
        if parsed.query or parsed.fragment:
            skipped_count += 1
            continue
        
        # Clean the path - remove trailing slash if exists
        clean_path = path.rstrip('/')
        
        # Split into segments
        segments = [segment for segment in clean_path.split('/') if segment]
        
        if len(segments) == 0:
            # Root domain - add trailing slash for consistency
            normalized_url = f"{parsed.scheme}://{parsed.netloc}/"
            filtered_urls.append(normalized_url)
            zero_segment_count += 1
            print(f"  ✓ [0 segments - ROOT] {url} -> normalized to: {normalized_url}")
        elif len(segments) == 1:
            # Single segment - add trailing slash for consistency
            normalized_url = f"{parsed.scheme}://{parsed.netloc}/{segments[0]}/"
            filtered_urls.append(normalized_url)
            one_segment_count += 1
            print(f"  ✓ [1 segment] {url} -> path: '{path}', segments: {segments}, normalized to: {normalized_url}")
        else:
            # 2+ segments - skip
            skipped_count += 1
            print(f"  ✗ [2+ segments] {url} -> path: '{path}', segments: {segments} (has {len(segments)} segments)")
    
    print(f"\n=== FILTERING SUMMARY ===")
    print(f"0 segments (root): {zero_segment_count}")
    print(f"1 segment: {one_segment_count}")
    print(f"2+ segments (skipped): {skipped_count}")
    print(f"Total filtered: {len(filtered_urls)} out of {len(urls)} URLs\n")
    
    return filtered_urls

# --- DEBUG FUNCTION ---
def debug_sitemap_urls(urls):
    """Debug function to see what URLs we're getting from sitemap"""
    print(f"\n=== SITEMAP URL SAMPLES (first 20) ===")
    for i, url in enumerate(urls[:20]):
        parsed = urlparse(url)
        clean_path = parsed.path.rstrip('/')
        segments = [s for s in clean_path.split('/') if s]
        print(f"{i+1:2d}. {url}")
        print(f"     Path: '{parsed.path}' | Segments: {segments} ({len(segments)} segments)")
    
    print(f"\n=== URL ANALYSIS ===")
    url_types = {}
    for url in urls:
        parsed = urlparse(url)
        clean_path = parsed.path.rstrip('/')
        segments = [s for s in clean_path.split('/') if s]
        seg_count = len(segments)
        url_types[seg_count] = url_types.get(seg_count, 0) + 1
    
    print(f"Segment distribution:")
    for seg_count in sorted(url_types.keys()):
        print(f"  {seg_count} segment(s): {url_types[seg_count]} URLs")

# --- EMAIL FUNCTION ---
def send_email(subject, message, recipients, attachment_path):
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ', '.join(recipients)
    msg['Subject'] = subject
    msg.attach(MIMEText(message, 'plain'))

    # Attach file
    try:
        with open(attachment_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(attachment_path)}"')
            msg.attach(part)
    except FileNotFoundError:
        print(f"Attachment file not found: {attachment_path}")
        return

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_username, smtp_password)
        text = msg.as_string()
        server.sendmail(sender_email, recipients, text)
        server.quit()
        print(f"Email sent to {', '.join(recipients)}")
    except Exception as e:
        print(f"Email to {', '.join(recipients)} failed with error: {str(e)}")

# --- SITEMAP URL FETCHING ---
def get_urls_from_sitemap(sitemap_url):
    """
    Downloads and parses a sitemap (or sitemap index) to extract all URLs.
    Handles both sitemap files, sitemap index files, and gzipped files recursively.
    """
    urls = set()
    sitemap_namespaces = {'sitemap': 'http://www.sitemaps.org/schemas/sitemap/0.9'}

    try:
        print(f"Fetching sitemap: {sitemap_url}")
        headers = {
            'User-Agent': 'Mozilla/5.0 (compatible; SEO-Monitor/1.0)'
        }
        response = requests.get(sitemap_url, headers=headers, timeout=30)
        response.raise_for_status()

        content_bytes = response.content

        # Handle compressed sitemaps (.gz)
        if sitemap_url.lower().endswith('.gz'):
            content = gzip.decompress(content_bytes).decode('utf-8')
            root = ET.fromstring(content)
        else:
            root = ET.fromstring(content_bytes)

        # Check if it's a SITEMAP INDEX (contains tags)
        sitemap_tags = root.findall('sitemap:sitemap', sitemap_namespaces)
        if sitemap_tags:
            print(f"Found sitemap index with {len(sitemap_tags)} child sitemaps. Recursing...")
            for sitemap_tag in sitemap_tags:
                child_sitemap_url = sitemap_tag.find('sitemap:loc', sitemap_namespaces)
                if child_sitemap_url is not None and child_sitemap_url.text:
                    urls.update(get_urls_from_sitemap(child_sitemap_url.text))

        # Process as a regular SITEMAP (contains tags)
        else:
            url_tags = root.findall('sitemap:url', sitemap_namespaces)
            if url_tags:
                print(f"Found regular sitemap with {len(url_tags)} URLs.")
                for url_tag in url_tags:
                    loc = url_tag.find('sitemap:loc', sitemap_namespaces)
                    if loc is not None and loc.text:
                        urls.add(loc.text)
            else:
                print(f"No URLs found in {sitemap_url} (or XML structure is unexpected).")

    except requests.exceptions.RequestException as e:
        print(f"Error fetching sitemap {sitemap_url}: {e}")
    except ET.ParseError as e:
        print(f"Error parsing XML for sitemap {sitemap_url}: {e}")
    except Exception as e:
        print(f"An unexpected error occurred while processing sitemap {sitemap_url}: {e}")

    return list(urls)

# --- PERFORMANCE MEASUREMENT FUNCTIONS ---
def get_cdp_metrics(driver):
    """Use Chrome DevTools Protocol to get metrics if possible."""
    try:
        metrics_payload = driver.execute_cdp_cmd('Performance.getMetrics', {})
        metrics = {item['name']: item['value'] for item in metrics_payload.get('metrics', [])}

        nav_start = metrics.get('NavigationStart', 0)
        fcp_metric = metrics.get('FirstContentfulPaint')
        load_metric = metrics.get('LoadEventEnd') or metrics.get('LoadEventStart')

        fcp_time = round((fcp_metric - nav_start) / 1000, 2) if fcp_metric is not None else None
        total_load_time = round((load_metric - nav_start) / 1000, 2) if load_metric is not None else None

        return fcp_time, total_load_time
    except Exception as e:
        return None, None

def process_url(url, base_domain):
    """
    Fetches FCP and Total Load Time using Selenium and performance APIs.
    """
    fcp_time = None
    total_load_time = None
    driver = None

    try:
        # First check if URL is accessible
        session = requests.Session()
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        }
        response = session.get(url, headers=headers, allow_redirects=True, timeout=30)
        response_code = response.status_code
        print(f"Processing: {url} - Status: {response_code}")

        # Setup Chrome driver
        chrome_options = Options()
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)

        # Initialize driver
        if WEBDRIVER_MANAGER_AVAILABLE:
            try:
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
            except Exception:
                driver = webdriver.Chrome(options=chrome_options)
        else:
            driver = webdriver.Chrome(options=chrome_options)

        if driver:
            # Enable CDP if available
            cdp_enabled = False
            if hasattr(driver, "execute_cdp_cmd"):
                try:
                    driver.execute_cdp_cmd('Performance.enable', {})
                    cdp_enabled = True
                except Exception:
                    pass

            # Load the page
            driver.get(url)
            time.sleep(3)  # Increased wait time for page to fully load

            # Try CDP metrics first
            if cdp_enabled:
                cdp_fcp, cdp_total = get_cdp_metrics(driver)
                fcp_time = cdp_fcp if cdp_fcp is not None else fcp_time
                total_load_time = cdp_total if cdp_total is not None else total_load_time

            # Fallback to JavaScript performance API
            if fcp_time is None or total_load_time is None:
                performance_script = """
                var perfData = window.performance.timing;
                var paintData = performance.getEntriesByType('paint');
                var result = {
                    navigationStart: perfData.navigationStart,
                    loadEventEnd: perfData.loadEventEnd,
                    fcp: null
                };
                for (var i = 0; i < paintData.length; i++) {
                    if (paintData[i].name === 'first-contentful-paint') {
                        result.fcp = paintData[i].startTime;
                        break;
                    }
                }
                return result;
                """
                perf_data = driver.execute_script(performance_script)

                if perf_data and perf_data.get('navigationStart'):
                    navigation_start = perf_data['navigationStart']
                    if total_load_time is None and perf_data.get('loadEventEnd'):
                        total_load_time = round((perf_data['loadEventEnd'] - navigation_start) / 1000, 2)
                    if fcp_time is None and perf_data.get('fcp'):
                        fcp_time = round(perf_data['fcp'] / 1000, 2)

            print(f" FCP: {fcp_time}s | Total: {total_load_time}s")

    except Exception as e:
        print(f"Error calculating page load metrics for {url}: {e}")

    finally:
        if driver:
            try:
                if cdp_enabled:
                    try:
                        driver.execute_cdp_cmd('Performance.disable', {})
                    except Exception:
                        pass
                driver.quit()
            except Exception as e:
                print(f"Error closing driver: {e}")

    seo_tags = {
        "URL": url,
        "First Contentful Paint (FCP) in seconds": fcp_time,
        "Total Page Load Time in seconds": total_load_time,
    }

    return seo_tags

def categorize_load_time(load_time):
    """Categorize Page Load Time into RAG status (Red/Amber/Green)"""
    if load_time is None:
        return "N/A"
    elif load_time < 3:
        return "Green"
    elif 3 <= load_time < 5:
        return "Amber"
    else:
        return "Red"

# --- REPORT GENERATION FUNCTIONS ---
def process_urls_from_sitemap(sitemap_url, project_name):
    """
    Fetches URLs from the sitemap, filters them, and processes them concurrently.
    """
    all_urls = get_urls_from_sitemap(sitemap_url)

    if not all_urls:
        print(f"No URLs found to process from sitemap: {sitemap_url}")
        return []

    # Debug: Show what we got from sitemap
    debug_sitemap_urls(all_urls)

    # Filter URLs
    urls = filter_urls_ending_with_slash(all_urls)

    if not urls:
        print(f"No URLs ending with '/' found in sitemap: {sitemap_url}")
        return []

    try:
        base_domain = urlparse(urls[0]).netloc
    except Exception:
        base_domain = "unknown"

    print(f"Starting performance check for {len(urls)} URLs from {sitemap_url} (Base Domain: {base_domain})")

    data = []
    # Reduced max_workers to avoid overwhelming the server
    with ThreadPoolExecutor(max_workers=5) as executor:
        results = executor.map(process_url, urls, [base_domain] * len(urls))
        for result in results:
            data.append(result)

    return data

def color_detailed_report(excel_file_path, load_rag_list):
    """Color only Load Time column based on RAG status"""
    try:
        book = load_workbook(excel_file_path)
        if 'Detailed Report' not in book.sheetnames:
            book.save(excel_file_path)
            return

        sheet = book['Detailed Report']

        fill_map = {
            "Green": PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
            "Amber": PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
            "Red": PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            "N/A": PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        }

        # Color Load Time column (column C)
        for row_index in range(len(load_rag_list)):
            excel_row = row_index + 2
            load_rag = load_rag_list[row_index]
            if load_rag in fill_map:
                sheet.cell(row=excel_row, column=3).fill = fill_map[load_rag]

        book.save(excel_file_path)
    except Exception as e:
        print(f"Error coloring detailed report: {e}")

def add_summary_sheet(excel_file_path, data):
    """Add a summary sheet with RAG analysis for Load Time only"""
    try:
        # Calculate RAG categories for Load Time with new thresholds
        load_green = sum(1 for item in data if categorize_load_time(item.get("Total Page Load Time in seconds")) == "Green")
        load_amber = sum(1 for item in data if categorize_load_time(item.get("Total Page Load Time in seconds")) == "Amber")
        load_red = sum(1 for item in data if categorize_load_time(item.get("Total Page Load Time in seconds")) == "Red")
        load_na = sum(1 for item in data if categorize_load_time(item.get("Total Page Load Time in seconds")) == "N/A")
        total = len(data)

        book = load_workbook(excel_file_path)
        if 'Summary' in book.sheetnames:
            del book['Summary']

        summary_sheet = book.create_sheet('Summary', 0)

        # Formatting styles
        green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        amber_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Title
        summary_sheet['A1'] = 'Page Load Time Performance Summary'
        summary_sheet['A1'].font = Font(bold=True, size=14)
        summary_sheet.merge_cells('A1:C1')
        summary_sheet['A1'].alignment = Alignment(horizontal='center')

        # Header Row
        summary_sheet['A3'] = 'RAG Category'
        summary_sheet['B3'] = 'Count'
        summary_sheet['C3'] = 'Percentage'

        for col in ['A3', 'B3', 'C3']:
            summary_sheet[col].fill = header_fill
            summary_sheet[col].font = header_font
            summary_sheet[col].alignment = Alignment(horizontal='center')
            summary_sheet[col].border = thin_border

        # Load Time Data Rows with updated thresholds
        load_data = [
            ('Green (< 3s)', load_green, green_fill),
            ('Amber (3s to < 5s)', load_amber, amber_fill),
            ('Red (>= 5s)', load_red, red_fill)
        ]

        row = 4
        for label, count, fill in load_data:
            summary_sheet[f'A{row}'] = label
            summary_sheet[f'B{row}'] = count
            summary_sheet[f'C{row}'] = f"{(count/total*100):.2f}%" if total > 0 else "0.00%"
            summary_sheet[f'A{row}'].fill = fill
            summary_sheet[f'B{row}'].alignment = Alignment(horizontal='center')
            summary_sheet[f'C{row}'].alignment = Alignment(horizontal='center')

            for col in ['A', 'B', 'C']:
                summary_sheet[f'{col}{row}'].border = thin_border
            row += 1

        summary_sheet[f'A{row}'] = f'N/A: {load_na}'
        summary_sheet[f'A{row}'].font = Font(italic=True)

        # Total Info
        row += 2
        summary_sheet[f'A{row}'] = f'Total URLs Analyzed: {total}'
        summary_sheet[f'A{row}'].font = Font(bold=True)

        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['C'].width = 15

        book.save(excel_file_path)
        print(f"Summary sheet added to {excel_file_path}")
    except Exception as e:
        print(f"Error adding summary sheet: {e}")

def process_and_save_reports_sitemap(sitemap_url, output_dir, project_name):
    """
    Main function to process URLs from Sitemap and save Excel reports.
    """
    data = process_urls_from_sitemap(sitemap_url, project_name)

    if not data:
        print(f"Skipping report generation for {project_name}: No data.")
        return []

    # Create DataFrame with only 3 columns
    df = pd.DataFrame(data)
    df = df[["URL", "First Contentful Paint (FCP) in seconds", "Total Page Load Time in seconds"]]

    # Sort by Load Time (descending)
    df = df.sort_values(by="Total Page Load Time in seconds", ascending=False, na_position='last')

    # Get RAG status list for coloring Load Time only
    load_rag_list = [categorize_load_time(row["Total Page Load Time in seconds"]) for _, row in df.iterrows()]

    excel_file_path = os.path.join(output_dir, f'{project_name}_performance_report_{timestamp}.xlsx')

    # Save to Excel
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Detailed Report', index=False)

    # Color and add summary
    color_detailed_report(excel_file_path, load_rag_list)
    add_summary_sheet(excel_file_path, data)

    print(f"Final report saved to {excel_file_path}")
    return data

# --- MAIN EXECUTION ---
if __name__ == '__main__':
    # Setup directories
    os.makedirs(output_dir_birla, exist_ok=True)

    start_time = time.time()

    # Process Birla Opus URLs from Sitemap
    print("\n--- Starting Birla Opus Performance Scan ---")
    data = process_and_save_reports_sitemap(birla_opus_sitemap_url, output_dir_birla, 'birla_opus')
    print("--- Birla Opus Scan Complete ---\n")

    end_time = time.time()
    execution_time = end_time - start_time
    execution_time_minutes = round(execution_time / 60, 2)

    print(f"Total script execution time: {execution_time_minutes} minutes")

    # Zip Reports
    zip_file_path = None
    if os.path.isdir(output_dir_birla) and os.listdir(output_dir_birla):
        zip_file_path = f'birla_opus_performance_report_{timestamp}.zip'
        with zipfile.ZipFile(zip_file_path, 'w') as zipf:
            for root, dirs, files in os.walk(output_dir_birla):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, output_dir_birla)
                    zipf.write(file_path, arcname=arcname)

    # Send Email
    if data and zip_file_path:
        # Calculate RAG summary for Load Time with new thresholds
        total_urls = len(data)
        load_green = sum(1 for item in data if categorize_load_time(item.get("Total Page Load Time in seconds")) == "Green")
        load_amber = sum(1 for item in data if categorize_load_time(item.get("Total Page Load Time in seconds")) == "Amber")
        load_red = sum(1 for item in data if categorize_load_time(item.get("Total Page Load Time in seconds")) == "Red")

        subject = f"Birla Opus Performance Report - {date_string}"
        body = f"Greetings, Birla Opus Team.\n\n"
        body += f"Kindly review the outcomes of a recent performance assessment for the website, "
        body += f"sourcing URLs directly from the sitemap: {birla_opus_sitemap_url}.\n\n"
        body += f"Report Generated: {date_string}\n\n"
        body += f"Note: This report only includes URLs with 0 or 1 path segment ending with '/'.\n"
        body += f"Examples: https://www.birlaopus.com/ and https://www.birlaopus.com/colour-catalogue/\n\n"
        body += f"Here is a summary of the findings:\n"
        body += f"Total URLs analyzed: {total_urls}\n\n"
        body += f"Page Load Time Summary:\n"
        body += f" - Green (< 3s): {load_green} ({load_green/total_urls*100:.1f}%)\n"
        body += f" - Amber (3s to < 5s): {load_amber} ({load_amber/total_urls*100:.1f}%)\n"
        body += f" - Red (>= 5s): {load_red} ({load_red/total_urls*100:.1f}%)\n\n"
        body += f"Detailed report attached in ZIP folder: '{os.path.basename(zip_file_path)}'\n"
        body += f"Execution time: {execution_time_minutes} minutes\n\n"
        body += f"Please feel free to review the attached report and let us know if you have any questions or concerns. "
        body += f"We appreciate your cooperation and support.\n\n"
        body += f"Thanks & Regards,\nQ.A Automation Team,\n'DEPT®'."

        send_email(subject, body, recipients_birla, zip_file_path)

    # Cleanup: Delete the temporary Excel files and folders
    def cleanup_reports(directory):
        if os.path.isdir(directory):
            for root, dirs, files in os.walk(directory, topdown=False):
                for file in files:
                    try:
                        os.remove(os.path.join(root, file))
                    except Exception as e:
                        print(f"Error removing file {file}: {e}")
                for dir in dirs:
                    try:
                        os.rmdir(os.path.join(root, dir))
                    except Exception as e:
                        print(f"Error removing directory {dir}: {e}")
            try:
                os.rmdir(directory)
                print(f"Cleaned up directory: {directory}")
            except Exception as e:
                print(f"Error removing main directory: {e}")

    cleanup_reports(output_dir_birla)

    # Delete the zip file after sending the email
    if zip_file_path and os.path.exists(zip_file_path):
        try:
            os.remove(zip_file_path)
            print(f"Deleted zip file: {zip_file_path}")
        except Exception as e:
            print(f"Error deleting zip file: {e}")

    print("\nCleanup complete.")
