import requests
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import multiprocessing
from functools import partial
import time
import xml.etree.ElementTree as ET

# Base domain and variables
unique_urls = set()
category = ['performance', 'best-practices', 'seo', 'accessibility']
today = date.today().strftime("%Y-%m-%d")
locale = 'en'
key = 'AIzaSyDy1hXmxjvsiaHqdANdVFFQldIhP52Si9Q'  
base_domain = None

# File extensions to ignore
ignored_extensions = ('.pdf', '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')

def fetch_urls_from_sitemap(sitemap_url):
    """Fetch all URLs from a sitemap XML"""
    try:
        print(f"Fetching sitemap: {sitemap_url}")
        
        # Add browser-like headers to avoid 406 errors
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }
        
        response = requests.get(sitemap_url, headers=headers, timeout=30)
        response.raise_for_status()
        
        # Parse XML
        root = ET.fromstring(response.content)
        
        # Handle sitemap namespace
        namespaces = {'ns': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
        
        urls = []
        
        # Check if this is a sitemap index (contains other sitemaps)
        sitemap_elements = root.findall('.//ns:sitemap/ns:loc', namespaces)
        
        if sitemap_elements:
            # This is a sitemap index, fetch URLs from each sitemap
            print(f"Found {len(sitemap_elements)} sub-sitemaps in index")
            for sitemap_loc in sitemap_elements:
                sub_sitemap_url = sitemap_loc.text
                print(f"Fetching URLs from sub-sitemap: {sub_sitemap_url}")
                urls.extend(fetch_urls_from_sitemap(sub_sitemap_url))
        else:
            # This is a regular sitemap, extract URLs
            url_elements = root.findall('.//ns:url/ns:loc', namespaces)
            
            for url_elem in url_elements:
                url = url_elem.text
                # Filter out ignored extensions
                if not url.lower().endswith(ignored_extensions):
                    urls.append(url)
                else:
                    print(f"Skipping (ignored extension): {url}")
            
            print(f"Found {len(urls)} valid URLs in this sitemap")
        
        return urls
        
    except requests.RequestException as e:
        print(f"Error fetching sitemap {sitemap_url}: {e}")
        return []
    except ET.ParseError as e:
        print(f"Error parsing sitemap XML {sitemap_url}: {e}")
        return []
    except Exception as e:
        print(f"Unexpected error with sitemap {sitemap_url}: {e}")
        return []

def webcorevitals(session, url, device, categories, today, key, locale):
    """Fetch web core vitals data from Google PageSpeed Insights API"""
    headers = {
        'Cache-Control': 'no-cache, no-store, must-revalidate',
        'Pragma': 'no-cache',
        'Expires': '0',
    }

    params = {
        'url': url,
        'key': key,
        'strategy': device,
        'locale': locale,
        'category': categories
    }

    try:
        response = session.get(
            "https://www.googleapis.com/pagespeedonline/v5/runPagespeed", 
            params=params, 
            headers=headers,
            timeout=60
        )
        response.raise_for_status()
        data = response.json()

        # Extract data
        loading_metrics = data.get('loadingExperience', {}).get('metrics', {})
        lighthouse_result = data.get('lighthouseResult', {})

        fid = loading_metrics.get("FIRST_INPUT_DELAY_MS", {}).get("percentile", 0)
        inp = loading_metrics.get("INTERACTION_TO_NEXT_PAINT", {}).get("percentile", 0)
        ttfb = loading_metrics.get("EXPERIMENTAL_TIME_TO_FIRST_BYTE", {}).get("percentile", 0)

        fcp = lighthouse_result.get('audits', {}).get('first-contentful-paint', {}).get('numericValue', 0)
        lcp = lighthouse_result.get('audits', {}).get('largest-contentful-paint', {}).get('numericValue', 0)
        cls = lighthouse_result.get('audits', {}).get('cumulative-layout-shift', {}).get('numericValue', 0)
        si = lighthouse_result.get('audits', {}).get('speed-index', {}).get('numericValue', 0)
        tti = lighthouse_result.get('audits', {}).get('interactive', {}).get('numericValue', 0)
        tbt = lighthouse_result.get('audits', {}).get('total-blocking-time', {}).get('numericValue', 0)
        bytes = lighthouse_result.get('audits', {}).get('total-byte-weight', {}).get('numericValue', 0)
        accessibility_score = lighthouse_result.get('categories', {}).get('accessibility', {}).get('score', 0) * 100
        score = lighthouse_result.get('categories', {}).get('performance', {}).get('score', 0) * 100
        best_practices_score = lighthouse_result.get('categories', {}).get('best-practices', {}).get('score', 0) * 100
        seo_score = lighthouse_result.get('categories', {}).get('seo', {}).get('score', 0) * 100

        # Create DataFrame
        df_score = pd.DataFrame({
            'URL': [url],
            'First Input Delay (ms)': [fid / 1000],
            'Interaction to Next Paint (ms)': [inp / 1000],
            'Time to First Byte (s)': [ttfb / 1000],
            'First Contentful Paint (ms)': [fcp / 1000],
            'Speed Index (ms)': [si],
            'Largest Contentful Paint (ms)': [lcp / 1000],
            'Time to Interactive (ms)': [tti / 1000],
            'Total Blocking Time (ms)': [tbt / 1000],
            'Cumulative Layout Shift': [cls],
            'Page Size (MB)': [bytes / (1024 * 1024)],
            'Date': [today],
            'Best Practices Score': [best_practices_score],
            'Accessibility Score': [accessibility_score],
            'Performance Score': [score],
            'SEO Score': [seo_score],
            'Device': [device]
        })

        print(f"✓ Successfully analyzed {url} ({device})")
        return df_score

    except requests.RequestException as e:
        print(f"✗ RequestException: Failed to fetch data for {url} ({device}) - {e}")
    except KeyError as e:
        print(f"✗ KeyError: {e}. No values found for URL: {url} ({device})")
    except Exception as e:
        print(f"✗ Error in webcorevitals for {url} ({device}): {e}")

    return pd.DataFrame()


def chunk_urls(urls, chunk_size):
    """Split URLs into chunks for batch processing"""
    return [urls[i:i + chunk_size] for i in range(0, len(urls), chunk_size)]


def run_analysis(urls):
    """Run Web Core Vitals analysis on all URLs"""
    session = requests.Session()
    chunk_size = 10  # Reduced chunk size to avoid rate limiting
    url_chunks = chunk_urls(urls, chunk_size)
    results = []

    print(f"\n{'='*60}")
    print(f"Starting analysis of {len(urls)} URLs in {len(url_chunks)} chunks")
    print(f"{'='*60}\n")

    for idx, chunk in enumerate(url_chunks, 1):
        print(f"\nProcessing chunk {idx}/{len(url_chunks)} ({len(chunk)} URLs)...")
        
        pool = multiprocessing.Pool(processes=4)  # Reduced processes to avoid rate limiting
        run_webcorevitals = partial(
            process_url, 
            session=session, 
            category=category, 
            today=today, 
            key=key, 
            locale=locale
        )
        chunk_results = pool.map(run_webcorevitals, chunk)
        pool.close()
        pool.join()
        results.extend(chunk_results)
        
        # Add delay between chunks to avoid rate limiting
        if idx < len(url_chunks):
            print(f"Waiting 5 seconds before next chunk...")
            time.sleep(5)

    # Check if there are valid data frames to concatenate
    valid_results = [result for result in results if not result.empty]
    if valid_results:
        concatenated_results = pd.concat(valid_results, ignore_index=True)
        print(f"\n{'='*60}")
        print(f"Analysis complete! Successfully analyzed {len(concatenated_results)} entries")
        print(f"{'='*60}\n")
        return concatenated_results
    else:
        print("\n✗ No valid DataFrames found. Cannot concatenate.")
        return pd.DataFrame()


def process_url(url, session, category, today, key, locale):
    """Process a single URL for both desktop and mobile"""
    print(f"\nProcessing: {url}")
    desktop_result = webcorevitals(session, url, 'desktop', category, today, key, locale)
    
    # Add small delay between desktop and mobile requests
    time.sleep(1)
    
    mobile_result = webcorevitals(session, url, 'mobile', category, today, key, locale)
    
    results = []
    if desktop_result is not None and not desktop_result.empty:
        results.append(desktop_result)
    if mobile_result is not None and not mobile_result.empty:
        results.append(mobile_result)
    
    if results:
        return pd.concat(results, ignore_index=True)
    else:
        return pd.DataFrame()


def apply_color_coding(worksheet, row_number):
    """Apply color coding to Excel cells based on performance thresholds"""
    for col_num in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row_number, column=col_num)
        value = cell.value
        column_name = worksheet.cell(row=1, column=col_num).value

        if value is None:
            continue
        
        try:
            if column_name == 'First Contentful Paint (ms)':
                value = float(value) / 1000
                if value <= 1.8:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 1.8 < value <= 3.0:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'Largest Contentful Paint (ms)':
                value = float(value) / 1000
                if value <= 2.5:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 2.5 < value <= 4.0:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'Cumulative Layout Shift':
                if value <= 0.1:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 0.1 < value <= 0.25:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'Total Blocking Time (ms)':
                value = float(value) / 1000
                if value <= 0.2:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 0.2 < value <= 0.6:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'First Input Delay (ms)':
                if value <= 100:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 100 < value <= 300:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'Page Size (MB)':
                if value <= 1:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 1 < value <= 3:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'Interaction to Next Paint (ms)':
                if value <= 200:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 200 < value <= 500:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'Speed Index (ms)':
                value = float(value) / 1000
                if value <= 3.4:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 3.4 < value <= 5.8:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'Time to Interactive (ms)':
                value = float(value) / 1000
                if value <= 3.8:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 3.8 < value <= 7.3:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'Time to First Byte (s)':
                if value <= 0.8:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 0.8 < value <= 1.8:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name == 'Performance Score':
                if value >= 90:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 50 <= value < 90:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            elif column_name in ['Best Practices Score', 'SEO Score', 'Accessibility Score']:
                if value >= 90:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif 50 <= value < 90:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        except Exception as e:
            print(f"Error applying color coding for {column_name} with value {value}: {e}")


def save_to_excel(df_results, directory):
    """Save analysis results to Excel with color coding"""
    timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"{directory}/web_core_vitals_results_{timestamp}.xlsx"
    
    # Separate desktop and mobile results
    df_desktop = df_results[df_results['Device'] == 'desktop'].drop(columns=['Device'])
    df_mobile = df_results[df_results['Device'] == 'mobile'].drop(columns=['Device'])

    # Create workbook
    wb = Workbook()

    # Create sheets for desktop and mobile results
    ws_desktop = wb.active
    ws_desktop.title = "Desktop Results"
    ws_mobile = wb.create_sheet(title="Mobile Results")

    # Write desktop results
    headers_desktop = df_desktop.columns.tolist()
    ws_desktop.append(headers_desktop)

    for r_idx, row in enumerate(df_desktop.itertuples(index=False, name=None), start=1):
        ws_desktop.append(row)
        if r_idx > 0:
            apply_color_coding(ws_desktop, r_idx + 1)

    # Write mobile results
    headers_mobile = df_mobile.columns.tolist()
    ws_mobile.append(headers_mobile)

    for r_idx, row in enumerate(df_mobile.itertuples(index=False, name=None), start=1):
        ws_mobile.append(row)
        if r_idx > 0:
            apply_color_coding(ws_mobile, r_idx + 1)

    # Save workbook
    wb.save(filename)
    print(f"\n✓ Results saved to: {filename}")
    print(f"  - Desktop results: {len(df_desktop)} URLs")
    print(f"  - Mobile results: {len(df_mobile)} URLs")


if __name__ == '__main__':
    print("\n" + "="*60)
    print("Web Core Vitals Analysis Tool")
    print("="*60)
    
    # Sitemap URL
    sitemap_url = 'https://www.birlaopus.com/sitemap.xml'
    
    print(f"\nSitemap URL: {sitemap_url}")
    print(f"Analysis Date: {today}")
    print("\n" + "-"*60)
    
    # Fetch URLs from sitemap
    print("\nStep 1: Fetching URLs from sitemap...")
    urls = fetch_urls_from_sitemap(sitemap_url)
    
    if not urls:
        print("\n✗ No URLs found in sitemap! Exiting...")
        exit(1)
    
    print(f"\n✓ Successfully fetched {len(urls)} URLs from sitemap")
    
    # Show first 5 URLs as preview
    print("\nFirst 5 URLs to be analyzed:")
    for i, url in enumerate(urls[:5], 1):
        print(f"  {i}. {url}")
    if len(urls) > 5:
        print(f"  ... and {len(urls) - 5} more")
    
    # Confirm before starting
    print(f"\n" + "-"*60)
    print("Step 2: Running Web Core Vitals analysis...")
    print("This may take a while depending on the number of URLs...")
    print("-"*60)
    
    # Run the analysis
    df_results = run_analysis(urls)
    
    if not df_results.empty:
        # Save the results to Excel
        print("\nStep 3: Saving results to Excel...")
        save_to_excel(df_results, directory='.')
        
        print("\n" + "="*60)
        print("✓ Web Core Vitals analysis completed successfully!")
        print("="*60 + "\n")
    else:
        print("\n" + "="*60)
        print("✗ Failed to analyze URLs - no valid results")
        print("="*60 + "\n")